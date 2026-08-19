from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
from urllib.parse import urlparse
from io import BytesIO
import sqlite3
import os
import uuid
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagemExcel
from openpyxl.styles import Font

app = Flask(__name__)
DB_PATH = os.path.join(os.path.dirname(__file__), "estoque.db")
PASTA_UPLOADS = os.path.join(os.path.dirname(__file__), "static", "uploads", "cotacoes")
os.makedirs(PASTA_UPLOADS, exist_ok=True)

EXTENSOES_PERMITIDAS = {"png", "jpg", "jpeg", "webp", "gif"}


# ---------- Banco de dados ----------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS estoques (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            criado_em TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            observacoes TEXT,
            estoque_id INTEGER NOT NULL,
            quantidade INTEGER NOT NULL DEFAULT 0,
            quantidade_minima INTEGER NOT NULL DEFAULT 0,
            preco REAL NOT NULL DEFAULT 0,
            criado_em TEXT,
            atualizado_em TEXT,
            FOREIGN KEY (estoque_id) REFERENCES estoques (id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            produto_nome TEXT NOT NULL,
            detalhes TEXT,
            data TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cotacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            criado_em TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS itens_cotacao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cotacao_id INTEGER NOT NULL,
            link TEXT,
            loja TEXT,
            nome TEXT NOT NULL,
            imagem_arquivo TEXT,
            preco REAL NOT NULL DEFAULT 0,
            frete REAL NOT NULL DEFAULT 0,
            observacao TEXT,
            criado_em TEXT,
            FOREIGN KEY (cotacao_id) REFERENCES cotacoes (id)
        )
    """)
    conn.commit()

    existentes = conn.execute("SELECT COUNT(*) as total FROM estoques").fetchone()
    if existentes["total"] == 0:
        agora = datetime.now().isoformat()
        for nome in ["Casa", "Sala de convívio", "Servidor"]:
            conn.execute("INSERT INTO estoques (nome, criado_em) VALUES (?, ?)", (nome, agora))
        conn.commit()

    conn.close()


def registrar_log(conn, tipo, produto_nome, detalhes=""):
    conn.execute(
        "INSERT INTO logs (tipo, produto_nome, detalhes, data) VALUES (?, ?, ?, ?)",
        (tipo, produto_nome, detalhes, datetime.now().isoformat())
    )


def buscar_nome_estoque(conn, estoque_id):
    linha = conn.execute("SELECT nome FROM estoques WHERE id = ?", (estoque_id,)).fetchone()
    return linha["nome"] if linha else "(estoque removido)"


# ---------- Páginas ----------

@app.route("/")
def pagina_estoque():
    return render_template("estoque.html")


@app.route("/gerente")
def pagina_gerente():
    return render_template("gerente.html")


@app.route("/logs")
def pagina_logs():
    return render_template("logs.html")


@app.route("/configuracoes")
def pagina_configuracoes():
    return render_template("configuracoes.html")


@app.route("/notificacoes")
def pagina_notificacoes():
    return render_template("notificacoes.html")


@app.route("/cotacoes")
def pagina_cotacoes():
    return render_template("cotacoes.html")


@app.route("/cotacoes/<int:cotacao_id>")
def pagina_cotacao_detalhe(cotacao_id):
    return render_template("cotacao_detalhe.html", cotacao_id=cotacao_id)


# ---------- API: estoques ----------

@app.route("/api/estoques", methods=["GET"])
def listar_estoques():
    conn = get_db()
    estoques = conn.execute("SELECT * FROM estoques ORDER BY nome").fetchall()
    conn.close()
    return jsonify([dict(e) for e in estoques])


@app.route("/api/estoques", methods=["POST"])
def criar_estoque():
    dados = request.get_json()
    nome = (dados.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "Nome do estoque é obrigatório"}), 400

    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO estoques (nome, criado_em) VALUES (?, ?)",
        (nome, datetime.now().isoformat())
    )
    registrar_log(conn, "estoque_criado", nome, "Novo estoque criado")
    conn.commit()
    estoque = conn.execute("SELECT * FROM estoques WHERE id = ?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(estoque)), 201


@app.route("/api/estoques/<int:estoque_id>", methods=["PUT"])
def editar_estoque(estoque_id):
    dados = request.get_json()
    nome = (dados.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "Nome do estoque é obrigatório"}), 400

    conn = get_db()
    estoque = conn.execute("SELECT * FROM estoques WHERE id = ?", (estoque_id,)).fetchone()
    if estoque is None:
        conn.close()
        return jsonify({"erro": "Estoque não encontrado"}), 404

    nome_antigo = estoque["nome"]
    conn.execute("UPDATE estoques SET nome = ? WHERE id = ?", (nome, estoque_id))
    if nome != nome_antigo:
        registrar_log(conn, "estoque_editado", nome, f"Nome alterado de '{nome_antigo}' para '{nome}'")
    conn.commit()
    estoque = conn.execute("SELECT * FROM estoques WHERE id = ?", (estoque_id,)).fetchone()
    conn.close()
    return jsonify(dict(estoque))


@app.route("/api/estoques/<int:estoque_id>", methods=["DELETE"])
def excluir_estoque(estoque_id):
    conn = get_db()
    estoque = conn.execute("SELECT * FROM estoques WHERE id = ?", (estoque_id,)).fetchone()
    if estoque is None:
        conn.close()
        return jsonify({"erro": "Estoque não encontrado"}), 404

    em_uso = conn.execute(
        "SELECT COUNT(*) as total FROM produtos WHERE estoque_id = ?", (estoque_id,)
    ).fetchone()["total"]

    if em_uso > 0:
        conn.close()
        return jsonify({
            "erro": f"Não é possível excluir: há {em_uso} produto(s) cadastrado(s) nesse estoque. Mova ou exclua esses produtos primeiro."
        }), 400

    conn.execute("DELETE FROM estoques WHERE id = ?", (estoque_id,))
    registrar_log(conn, "estoque_excluido", estoque["nome"], "Estoque removido do sistema")
    conn.commit()
    conn.close()
    return jsonify({"sucesso": True})


# ---------- API: produtos ----------

@app.route("/api/produtos", methods=["GET"])
def listar_produtos():
    conn = get_db()
    produtos = conn.execute("""
        SELECT produtos.*, estoques.nome AS estoque_nome
        FROM produtos
        JOIN estoques ON estoques.id = produtos.estoque_id
        ORDER BY produtos.nome
    """).fetchall()
    conn.close()
    return jsonify([dict(p) for p in produtos])


@app.route("/api/produtos", methods=["POST"])
def criar_produto():
    dados = request.get_json()

    nome = (dados.get("nome") or "").strip()
    if not nome:
        return jsonify({"erro": "Nome do produto é obrigatório"}), 400

    estoque_id = dados.get("estoque_id")
    if not estoque_id:
        return jsonify({"erro": "Selecione um estoque"}), 400

    observacoes = (dados.get("observacoes") or "").strip()
    quantidade = int(dados.get("quantidade", 0))
    quantidade_minima = int(dados.get("quantidade_minima", 0))

    agora = datetime.now().isoformat()
    conn = get_db()

    nome_estoque = buscar_nome_estoque(conn, estoque_id)

    cursor = conn.execute(
        "INSERT INTO produtos (nome, observacoes, estoque_id, quantidade, quantidade_minima, criado_em, atualizado_em) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (nome, observacoes, estoque_id, quantidade, quantidade_minima, agora, agora)
    )
    novo_id = cursor.lastrowid

    detalhes = f"Cadastrado em '{nome_estoque}' com estoque mínimo {quantidade_minima}"
    if observacoes:
        detalhes += f" — Obs: {observacoes}"
    registrar_log(conn, "criacao", nome, detalhes)

    conn.commit()
    produto = conn.execute("""
        SELECT produtos.*, estoques.nome AS estoque_nome
        FROM produtos JOIN estoques ON estoques.id = produtos.estoque_id
        WHERE produtos.id = ?
    """, (novo_id,)).fetchone()
    conn.close()

    return jsonify(dict(produto)), 201


@app.route("/api/produtos/<int:produto_id>", methods=["PUT"])
def editar_produto(produto_id):
    dados = request.get_json()
    conn = get_db()
    produto = conn.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if produto is None:
        conn.close()
        return jsonify({"erro": "Produto não encontrado"}), 404

    nome = dados.get("nome", produto["nome"])
    observacoes = dados.get("observacoes", produto["observacoes"])
    estoque_id = dados.get("estoque_id", produto["estoque_id"])
    quantidade_minima = int(dados.get("quantidade_minima", produto["quantidade_minima"]))

    mudancas = []
    if nome != produto["nome"]:
        mudancas.append(f"nome '{produto['nome']}' → '{nome}'")
    if observacoes != (produto["observacoes"] or ""):
        antiga = produto["observacoes"] or "(vazio)"
        nova = observacoes or "(vazio)"
        mudancas.append(f"observações: '{antiga}' → '{nova}'")
    if estoque_id != produto["estoque_id"]:
        nome_antigo = buscar_nome_estoque(conn, produto["estoque_id"])
        nome_novo = buscar_nome_estoque(conn, estoque_id)
        mudancas.append(f"estoque '{nome_antigo}' → '{nome_novo}'")
    if quantidade_minima != produto["quantidade_minima"]:
        mudancas.append(f"mínimo {produto['quantidade_minima']} → {quantidade_minima}")
    detalhes = "; ".join(mudancas) if mudancas else "Nenhuma alteração detectada"

    conn.execute(
        "UPDATE produtos SET nome = ?, observacoes = ?, estoque_id = ?, quantidade_minima = ?, atualizado_em = ? WHERE id = ?",
        (nome, observacoes, estoque_id, quantidade_minima, datetime.now().isoformat(), produto_id)
    )
    registrar_log(conn, "edicao", nome, detalhes)

    conn.commit()
    produto = conn.execute("""
        SELECT produtos.*, estoques.nome AS estoque_nome
        FROM produtos JOIN estoques ON estoques.id = produtos.estoque_id
        WHERE produtos.id = ?
    """, (produto_id,)).fetchone()
    conn.close()
    return jsonify(dict(produto))


@app.route("/api/produtos/<int:produto_id>", methods=["DELETE"])
def excluir_produto(produto_id):
    conn = get_db()
    produto = conn.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if produto is None:
        conn.close()
        return jsonify({"erro": "Produto não encontrado"}), 404

    conn.execute("DELETE FROM produtos WHERE id = ?", (produto_id,))
    registrar_log(conn, "exclusao", produto["nome"], "Produto removido do sistema")
    conn.commit()
    conn.close()
    return jsonify({"sucesso": True})


@app.route("/api/produtos/<int:produto_id>/movimentar", methods=["POST"])
def movimentar_estoque(produto_id):
    dados = request.get_json()
    tipo = dados.get("tipo")
    quantidade = int(dados.get("quantidade", 0))
    observacao = (dados.get("observacao") or "").strip()

    if tipo not in ("entrada", "saida"):
        return jsonify({"erro": "Tipo deve ser 'entrada' ou 'saida'"}), 400
    if quantidade <= 0:
        return jsonify({"erro": "Quantidade deve ser maior que zero"}), 400

    conn = get_db()
    produto = conn.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if produto is None:
        conn.close()
        return jsonify({"erro": "Produto não encontrado"}), 404

    nova_quantidade = produto["quantidade"] + quantidade if tipo == "entrada" else produto["quantidade"] - quantidade
    if nova_quantidade < 0:
        conn.close()
        return jsonify({"erro": "Estoque insuficiente para essa saída"}), 400

    agora = datetime.now().isoformat()
    conn.execute("UPDATE produtos SET quantidade = ?, atualizado_em = ? WHERE id = ?", (nova_quantidade, agora, produto_id))

    detalhes = f"{quantidade} un. — Motivo: {observacao}" if observacao else f"{quantidade} un. — Motivo não informado"
    registrar_log(conn, tipo, produto["nome"], detalhes)

    conn.commit()
    produto_atualizado = conn.execute("""
        SELECT produtos.*, estoques.nome AS estoque_nome
        FROM produtos JOIN estoques ON estoques.id = produtos.estoque_id
        WHERE produtos.id = ?
    """, (produto_id,)).fetchone()
    conn.close()
    return jsonify(dict(produto_atualizado))


@app.route("/api/logs", methods=["GET"])
def listar_logs():
    limite = int(request.args.get("limite", 30))
    tipos_param = request.args.get("tipos")

    conn = get_db()
    if tipos_param:
        tipos = tipos_param.split(",")
        marcadores = ",".join(["?"] * len(tipos))
        linhas = conn.execute(
            f"SELECT * FROM logs WHERE tipo IN ({marcadores}) ORDER BY data DESC LIMIT ?",
            (*tipos, limite)
        ).fetchall()
    else:
        linhas = conn.execute("SELECT * FROM logs ORDER BY data DESC LIMIT ?", (limite,)).fetchall()
    conn.close()
    return jsonify([dict(l) for l in linhas])


# ============================================================
# COTAÇÕES
# ============================================================

def extensao_valida(nome_arquivo):
    return "." in nome_arquivo and nome_arquivo.rsplit(".", 1)[1].lower() in EXTENSOES_PERMITIDAS


def salvar_arquivo_imagem(arquivo_upload):
    """Salva um arquivo de imagem enviado pelo formulário e devolve o nome salvo."""
    if not arquivo_upload or arquivo_upload.filename == "":
        return None
    if not extensao_valida(arquivo_upload.filename):
        return None
    extensao = arquivo_upload.filename.rsplit(".", 1)[1].lower()
    nome_salvo = f"{uuid.uuid4().hex}.{extensao}"
    arquivo_upload.save(os.path.join(PASTA_UPLOADS, nome_salvo))
    return nome_salvo


LOJAS_CONHECIDAS = {
    "mercadolivre.com.br": "Mercado Livre",
    "magazineluiza.com.br": "Magazine Luiza",
    "magazinevoce.com.br": "Magazine Luiza",
    "amazon.com.br": "Amazon",
    "americanas.com.br": "Americanas",
    "casasbahia.com.br": "Casas Bahia",
    "pontofrio.com.br": "Ponto Frio",
    "extra.com.br": "Extra",
    "shopee.com.br": "Shopee",
    "aliexpress.com": "AliExpress",
    "kabum.com.br": "KaBuM!",
    "pichau.com.br": "Pichau",
    "submarino.com.br": "Submarino",
    "netshoes.com.br": "Netshoes",
    "centauro.com.br": "Centauro",
    "carrefour.com.br": "Carrefour",
    "fastshop.com.br": "Fast Shop",
}


def formatar_nome_loja(dominio):
    """Deixa o nome da loja apresentável quando só temos o domínio
    (ex: 'mercadolivre.com.br' -> 'Mercado Livre'). Lojas conhecidas usam
    o nome oficial; as demais ganham pelo menos capitalização decente."""
    if not dominio:
        return dominio

    dominio_limpo = dominio.lower().replace("www.", "")
    if dominio_limpo in LOJAS_CONHECIDAS:
        return LOJAS_CONHECIDAS[dominio_limpo]

    primeira_parte = dominio_limpo.split(".")[0]
    return primeira_parte.replace("-", " ").replace("_", " ").title()


def baixar_imagem_da_url(imagem_url, headers):
    """Baixa uma imagem de uma URL externa e salva localmente. Devolve o nome salvo ou None."""
    try:
        resposta = requests.get(imagem_url, headers=headers, timeout=8)
        resposta.raise_for_status()
        content_type = resposta.headers.get("Content-Type", "")
        if "png" in content_type:
            extensao = "png"
        elif "webp" in content_type:
            extensao = "webp"
        elif "gif" in content_type:
            extensao = "gif"
        else:
            extensao = "jpg"
        nome_salvo = f"{uuid.uuid4().hex}.{extensao}"
        with open(os.path.join(PASTA_UPLOADS, nome_salvo), "wb") as f:
            f.write(resposta.content)
        return nome_salvo
    except Exception:
        return None


@app.route("/api/buscar-produto", methods=["POST"])
def buscar_produto():
    dados = request.get_json()
    url = (dados.get("url") or "").strip()
    if not url:
        return jsonify({"erro": "Link é obrigatório"}), 400

    resultado = {"loja": None, "nome": None, "imagem_arquivo": None, "erro": None}

    # A loja, na pior das hipóteses, sempre dá pra tirar do próprio link
    # (não depende de conseguir acessar o site)
    try:
        dominio = urlparse(url).netloc.replace("www.", "")
        resultado["loja"] = formatar_nome_loja(dominio)
    except Exception:
        pass

    headers = {"User-Agent": "Mozilla/5.0 (compatible; SistemaEstoqueBot/1.0)"}
    try:
        resposta = requests.get(url, headers=headers, timeout=8)
        resposta.raise_for_status()
        soup = BeautifulSoup(resposta.text, "html.parser")

        def meta(propriedade):
            tag = soup.find("meta", property=propriedade)
            return tag["content"].strip() if tag and tag.get("content") else None

        nome = meta("og:title")
        if nome:
            resultado["nome"] = nome

        site_name = meta("og:site_name")
        if site_name:
            resultado["loja"] = site_name

        imagem_url = meta("og:image")
        if imagem_url:
            resultado["imagem_arquivo"] = baixar_imagem_da_url(imagem_url, headers)

        if not resultado["nome"] and not resultado["imagem_arquivo"]:
            resultado["erro"] = "Não conseguimos identificar automaticamente os dados desse link"

    except Exception:
        resultado["erro"] = "Não foi possível acessar esse link automaticamente"

    return jsonify(resultado)


@app.route("/api/cotacoes", methods=["GET"])
def listar_cotacoes():
    conn = get_db()
    cotacoes = conn.execute("""
        SELECT c.*,
               COUNT(i.id) as total_itens,
               COALESCE(SUM(i.preco + i.frete), 0) as valor_total
        FROM cotacoes c
        LEFT JOIN itens_cotacao i ON i.cotacao_id = c.id
        GROUP BY c.id
        ORDER BY c.criado_em DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(c) for c in cotacoes])


@app.route("/api/cotacoes", methods=["POST"])
def criar_cotacao():
    dados = request.get_json()
    titulo = (dados.get("titulo") or "").strip()
    if not titulo:
        return jsonify({"erro": "Título é obrigatório"}), 400

    conn = get_db()
    cursor = conn.execute(
        "INSERT INTO cotacoes (titulo, criado_em) VALUES (?, ?)",
        (titulo, datetime.now().isoformat())
    )
    conn.commit()
    cotacao = conn.execute("SELECT * FROM cotacoes WHERE id = ?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(cotacao)), 201


@app.route("/api/cotacoes/<int:cotacao_id>", methods=["GET"])
def obter_cotacao(cotacao_id):
    conn = get_db()
    cotacao = conn.execute("SELECT * FROM cotacoes WHERE id = ?", (cotacao_id,)).fetchone()
    if cotacao is None:
        conn.close()
        return jsonify({"erro": "Cotação não encontrada"}), 404

    itens = conn.execute(
        "SELECT * FROM itens_cotacao WHERE cotacao_id = ? ORDER BY id", (cotacao_id,)
    ).fetchall()
    conn.close()

    resultado = dict(cotacao)
    resultado["itens"] = [dict(i) for i in itens]
    return jsonify(resultado)


@app.route("/api/cotacoes/<int:cotacao_id>", methods=["DELETE"])
def excluir_cotacao(cotacao_id):
    conn = get_db()
    cotacao = conn.execute("SELECT * FROM cotacoes WHERE id = ?", (cotacao_id,)).fetchone()
    if cotacao is None:
        conn.close()
        return jsonify({"erro": "Cotação não encontrada"}), 404

    itens = conn.execute("SELECT * FROM itens_cotacao WHERE cotacao_id = ?", (cotacao_id,)).fetchall()
    for item in itens:
        if item["imagem_arquivo"]:
            caminho = os.path.join(PASTA_UPLOADS, item["imagem_arquivo"])
            if os.path.exists(caminho):
                os.remove(caminho)

    conn.execute("DELETE FROM itens_cotacao WHERE cotacao_id = ?", (cotacao_id,))
    conn.execute("DELETE FROM cotacoes WHERE id = ?", (cotacao_id,))
    conn.commit()
    conn.close()
    return jsonify({"sucesso": True})


@app.route("/api/cotacoes/<int:cotacao_id>/itens", methods=["POST"])
def criar_item_cotacao(cotacao_id):
    conn = get_db()
    cotacao = conn.execute("SELECT * FROM cotacoes WHERE id = ?", (cotacao_id,)).fetchone()
    if cotacao is None:
        conn.close()
        return jsonify({"erro": "Cotação não encontrada"}), 404

    dados = request.form
    nome = (dados.get("nome") or "").strip()
    if not nome:
        conn.close()
        return jsonify({"erro": "Nome do produto é obrigatório"}), 400

    link = (dados.get("link") or "").strip() or None
    loja = (dados.get("loja") or "").strip()
    preco = float(dados.get("preco") or 0)
    frete = float(dados.get("frete") or 0)
    observacao = (dados.get("observacao") or "").strip()

    imagem_arquivo = salvar_arquivo_imagem(request.files.get("imagem"))
    if not imagem_arquivo:
        imagem_arquivo = (dados.get("imagem_arquivo_existente") or "").strip() or None

    cursor = conn.execute(
        """INSERT INTO itens_cotacao
           (cotacao_id, link, loja, nome, imagem_arquivo, preco, frete, observacao, criado_em)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (cotacao_id, link, loja, nome, imagem_arquivo, preco, frete, observacao, datetime.now().isoformat())
    )
    conn.commit()
    item = conn.execute("SELECT * FROM itens_cotacao WHERE id = ?", (cursor.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(item)), 201


@app.route("/api/cotacoes/<int:cotacao_id>/itens/<int:item_id>", methods=["PUT"])
def editar_item_cotacao(cotacao_id, item_id):
    conn = get_db()
    item = conn.execute(
        "SELECT * FROM itens_cotacao WHERE id = ? AND cotacao_id = ?", (item_id, cotacao_id)
    ).fetchone()
    if item is None:
        conn.close()
        return jsonify({"erro": "Item não encontrado"}), 404

    dados = request.form
    nome = (dados.get("nome") or "").strip()
    if not nome:
        conn.close()
        return jsonify({"erro": "Nome do produto é obrigatório"}), 400

    link = (dados.get("link") or "").strip() or None
    loja = (dados.get("loja") or "").strip()
    preco = float(dados.get("preco") or 0)
    frete = float(dados.get("frete") or 0)
    observacao = (dados.get("observacao") or "").strip()

    # Se uma imagem nova foi enviada, troca (e apaga o arquivo antigo).
    # Senão, mantém a imagem que o item já tinha.
    novo_arquivo = salvar_arquivo_imagem(request.files.get("imagem"))
    if novo_arquivo:
        if item["imagem_arquivo"]:
            caminho_antigo = os.path.join(PASTA_UPLOADS, item["imagem_arquivo"])
            if os.path.exists(caminho_antigo):
                os.remove(caminho_antigo)
        imagem_arquivo = novo_arquivo
    else:
        imagem_arquivo = item["imagem_arquivo"]

    conn.execute(
        """UPDATE itens_cotacao
           SET link = ?, loja = ?, nome = ?, imagem_arquivo = ?, preco = ?, frete = ?, observacao = ?
           WHERE id = ?""",
        (link, loja, nome, imagem_arquivo, preco, frete, observacao, item_id)
    )
    conn.commit()
    item_atualizado = conn.execute("SELECT * FROM itens_cotacao WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    return jsonify(dict(item_atualizado))


@app.route("/api/cotacoes/<int:cotacao_id>/itens/<int:item_id>", methods=["DELETE"])
def excluir_item_cotacao(cotacao_id, item_id):
    conn = get_db()
    item = conn.execute(
        "SELECT * FROM itens_cotacao WHERE id = ? AND cotacao_id = ?", (item_id, cotacao_id)
    ).fetchone()
    if item is None:
        conn.close()
        return jsonify({"erro": "Item não encontrado"}), 404

    if item["imagem_arquivo"]:
        caminho = os.path.join(PASTA_UPLOADS, item["imagem_arquivo"])
        if os.path.exists(caminho):
            os.remove(caminho)

    conn.execute("DELETE FROM itens_cotacao WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"sucesso": True})


@app.route("/api/cotacoes/<int:cotacao_id>/exportar", methods=["GET"])
def exportar_cotacao(cotacao_id):
    conn = get_db()
    cotacao = conn.execute("SELECT * FROM cotacoes WHERE id = ?", (cotacao_id,)).fetchone()
    if cotacao is None:
        conn.close()
        return jsonify({"erro": "Cotação não encontrada"}), 404

    itens = conn.execute(
        "SELECT * FROM itens_cotacao WHERE cotacao_id = ? ORDER BY id", (cotacao_id,)
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotação"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 34

    fonte_titulo = Font(bold=True, size=13)
    fonte_produto = Font(bold=True, size=11)
    fonte_cabecalho = Font(bold=True)
    FORMATO_MOEDA = 'R$ #,##0.00'

    ws.merge_cells("A1:D1")
    ws["A1"] = f"COTAÇÃO {cotacao['titulo'].upper()}"
    ws["A1"].font = fonte_titulo

    linha = 3
    ESPACO_POR_BLOCO = 13  # linhas reservadas por item, pra caber a foto sem sobrepor o próximo

    for item in itens:
        tem_observacao = bool(item["observacao"] and item["observacao"].strip())

        # Nome do produto pesquisado (identifica o bloco)
        ws.merge_cells(f"A{linha}:D{linha}")
        ws.cell(row=linha, column=1, value=item["nome"]).font = fonte_produto
        linha += 1

        # Cabeçalho da mini-tabela — "Observação" só entra se o item tiver uma
        cabecalhos = ["LOJA", "VALOR", "FRETE", "TOTAL", "Link"]
        if tem_observacao:
            cabecalhos.append("Observação")
        for col, texto in enumerate(cabecalhos, start=1):
            ws.cell(row=linha, column=col, value=texto).font = fonte_cabecalho

        linha_dados = linha + 1
        ws.cell(row=linha_dados, column=1, value=item["loja"] or "")

        celula_valor = ws.cell(row=linha_dados, column=2, value=item["preco"])
        celula_valor.number_format = FORMATO_MOEDA

        celula_frete = ws.cell(row=linha_dados, column=3, value=item["frete"])
        celula_frete.number_format = FORMATO_MOEDA

        celula_total = ws.cell(row=linha_dados, column=4, value=f"=SUM(B{linha_dados}:C{linha_dados})")
        celula_total.number_format = FORMATO_MOEDA

        if item["link"]:
            celula_link = ws.cell(row=linha_dados, column=5, value="Link")
            celula_link.hyperlink = item["link"]
            celula_link.font = Font(color="0563C1", underline="single")

        if tem_observacao:
            ws.cell(row=linha_dados, column=6, value=item["observacao"])

        if item["imagem_arquivo"]:
            caminho_imagem = os.path.join(PASTA_UPLOADS, item["imagem_arquivo"])
            if os.path.exists(caminho_imagem):
                try:
                    img = ImagemExcel(caminho_imagem)
                    img.width = 220
                    img.height = 220
                    ws.add_image(img, f"A{linha_dados + 1}")
                except Exception:
                    pass

        linha += ESPACO_POR_BLOCO

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo_seguro = "".join(
        c if c.isalnum() or c in (" ", "-", "_") else "_" for c in cotacao["titulo"]
    ).strip().replace(" ", "_")

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"cotacao_{nome_arquivo_seguro}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)